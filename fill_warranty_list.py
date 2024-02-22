from openpyxl import load_workbook
from datetime import datetime

def fill_warranty_list(customer_name, customer_number, tire_brand, tire_model, tire_size, tire_season, check_sum, NFD, check_number, manager, unique_code):
    # открываем список бшм и делаем запись в последнюю пустую строку

    time_stamp = datetime.now()
    form_data = [customer_name, customer_number, tire_brand, tire_model,
                 tire_size, tire_season, check_sum, NFD, check_number, manager, time_stamp, unique_code]

    wb_warranty_list = load_workbook(filename='base/warranty_list.xlsx')
    wb_warranty_list_sheet = wb_warranty_list.worksheets[0]

    for row in wb_warranty_list_sheet.iter_rows(min_row=1, min_col=1, max_col=12):
        step = 0
        for cell in row:
            if cell.offset(row=1).value == None:
                cell.offset(row=1).value = form_data[step]
            step += 1

    wb_warranty_list.save(filename='base/warranty_list.xlsx')



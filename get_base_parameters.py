from openpyxl import load_workbook

def get_base_parameters():

    wb_base_list = load_workbook(filename='base/base.xlsx')
    wb_base_list_sheet = wb_base_list['managers']

    managers = []
    tire_size_width = []
    tire_size_profile = []
    tire_size_diameter = []
    tire_brand = []
    tire_diameter_conditions = []

    for row in wb_base_list_sheet.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            managers.append(cell.value)

    wb_base_list_sheet = wb_base_list['sizes']

    for row in wb_base_list_sheet.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value != None:
                tire_size_width.append(cell.value)
            else:
                break

    for row in wb_base_list_sheet.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            if cell.value != None:
                tire_size_profile.append(cell.value)
            else:
                break

    for row in wb_base_list_sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if cell.value != None:
                tire_size_diameter.append(cell.value)
            else:
                break

    wb_base_list_sheet = wb_base_list['model_conditions']

    for row in wb_base_list_sheet.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value != None:
                tire_brand.append(cell.value)
            else:
                break

    brands_dict = {}
    models = []
    brand_name = ''

    wb_base_list_sheet = wb_base_list['model_conditions']

    for row in wb_base_list_sheet.iter_rows(min_row=2, min_col=2):
        step = 0
        for cell in row:
            if step == 0:
                brand_name = cell.offset(column=-1).value
                brands_dict[cell.offset(column=-1).value] = []
                step += 1
            if cell.value != None:
                models.append(cell.value)
            else:
                break
        brands_dict[brand_name] = models
        models = []

    wb_base_list_sheet = wb_base_list['compensation_conditions']

    compensations_dict = {}
    compensation_sum = {}

    for row in wb_base_list_sheet.iter_rows(min_row=2, min_col=2):
        step = 0
        compensation_diam = 13
        for cell in row:
            if step == 0:
                brand_name = cell.offset(column=-1).value
                compensations_dict[cell.offset(column=-1).value] = []
                step += 1
            if cell.value != None:
                compensation_sum[f'R{compensation_diam}'] = cell.value
                compensation_diam += 1
            else:
                break

        compensations_dict[brand_name] = compensation_sum.copy()
        compensation_sum.clear()

    wb_base_list_sheet = wb_base_list['budgets']

    budgets_dict = {}
    budgets_sum = {}

    for row in wb_base_list_sheet.iter_rows(min_row=2, min_col=2):
        step = 0
        for cell in row:
            if step == 0:
                brand_name = cell.offset(column=-1).value
                budgets_dict[cell.offset(column=-1).value] = []
                step += 1
            if cell.value != None:
                budgets_sum[f'budgeted'] = cell.value
                budgets_sum[f'budget_balance'] = cell.offset(column=1).value
                break

        budgets_dict[brand_name] = budgets_sum.copy()
        budgets_sum.clear()

    return managers, tire_size_width, tire_size_profile, tire_size_diameter, tire_brand, brands_dict, compensations_dict, budgets_dict

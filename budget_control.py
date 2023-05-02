from openpyxl import load_workbook
def budget_control(tire_brand, tire_size_diameter, brands_budgets, compensation_dict):
    new_budget = brands_budgets[tire_brand]['budget_balance'] - compensation_dict[tire_brand][tire_size_diameter]

    wb_warranty_list = load_workbook(filename='base/base.xlsx')
    wb_warranty_list_sheet = wb_warranty_list['budgets']

    for row in wb_warranty_list_sheet.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            if cell.value == tire_brand:
                cell.offset(column=2).value = int(new_budget)

    wb_warranty_list.save(filename='base/base.xlsx')

